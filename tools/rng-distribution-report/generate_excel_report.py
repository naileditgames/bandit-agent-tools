import json
import os
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

COLOR_TITLE_BG = "6D9EEB"
COLOR_TITLE_TEXT = "FFFFFF"
COLOR_SUBTITLE_BG = "A4C2F4"
COLOR_SUBTITLE_TEXT = "434343"

class ExcelStatsProcessor:
    def __init__(self):
        self._wb = Workbook()
        self._data = None
        self._skip = []
        self._skip_limit = -1;

        # remove default sheet
        if 'Sheet' in self._wb.sheetnames:
            self._wb.remove(self._wb['Sheet'])

    def skip_context(self, context):
        self._skip.append(context)

    def set_skip_limit(self, limit):
        self._skip_limit = limit

    def create_excel_report(self, data):
        self._data = data
        self._create_summary()
        self._create_variants_stats()

    def _create_summary(self):
        ws = self._wb.create_sheet(title="SUMMARY")
        start_column = 2
        used_columns = 3
        margin = 1

        for variant, variant_stats in self._data["stats"]["statsPerVariant"].items():
            self._print_single_variant_summary(ws, variant, variant_stats, start_column)
            start_column = start_column + used_columns + margin

        # change default widhh
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = 18

    def _print_single_variant_summary(self, ws, variant_name, variant_stats, start_column):
        current_column = start_column
        start_row = 2
        length = 0

        # general information
        ws.merge_cells(start_row=start_row, start_column=start_column, end_row=start_row,end_column=start_column + 2)
        self._add_header(ws, row=start_row, column=start_column, value="General Information")

        self._add_sub_header(ws, row=start_row + 1, column=start_column, value="variant")
        ws.merge_cells(start_row=start_row + 1, end_row=start_row + 1, start_column=start_column + 1, end_column=start_column + 2)
        self._add_text(ws, row=start_row + 1, column=start_column + 1, value=variant_name)

        self._add_sub_header(ws, row=start_row + 2, column=start_column, value="contexts")
        ws.merge_cells(start_row=start_row + 2, end_row=start_row + 2, start_column=start_column + 1, end_column=start_column + 2)
        self._add_text(ws, row=start_row + 2, column=start_column + 1, value=variant_stats["uniqueContexts"])

        self._add_sub_header(ws, row=start_row + 3, column=start_column, value="weight sets")
        ws.merge_cells(start_row=start_row + 3, end_row=start_row + 3, start_column=start_column + 1, end_column=start_column + 2)
        self._add_text(ws, row=start_row + 3, column=start_column + 1, value=variant_stats["uniqueWeightSets"])

        length = length + 4

        # weight sets per context
        hit_table_start_row = start_row + 4
        ws.merge_cells(start_row=hit_table_start_row, start_column=start_column, end_row=hit_table_start_row, end_column=start_column + 2)
        self._add_header(ws, row=hit_table_start_row, column=start_column, value="Weight Sets Per Context")

        self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 0, value="context")
        self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 1, value="weightSets")
        self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 2, value="skipped")

        for index, (context, hits) in enumerate(variant_stats["weightSetsPerContext"].items()):
            self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 0, value=context)
            self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 1, value=hits)
            self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 2, value=("SKIP" if (context in self._skip or hits > self._skip_limit) else ""))

        length = length + 2 + len(variant_stats["weightSetsPerContext"])

        # create borders
        thin = Side(style='thin')
        thick = Side(style='thick')
        for row in range(start_row, start_row + length - 1+ 1):
            for col in range(start_column, start_column + 2 + 1):
                cell = ws.cell(row=row, column=col)
                border = Border(
                    left=thick if col == start_column else thin,
                    right=thick if col == start_column + 2 else thin,
                    top=thick if row == start_row else thin,
                    bottom=thick if row == start_row + length - 1 else thin
                )
                cell.border = border

        return length

    def _create_variants_stats(self):
        for key, value in self._data["rng_distribution"]["variants"].items():
            self._create_single_variant_page(key, value)

    def _create_single_variant_page(self, variant_name, variant_stats):
        print(f"processing variant: {variant_name}")
        ws = self._wb.create_sheet(title=variant_name)
        start_column = 2
        used_columns = 3
        margin = 1

        for context, context_stats in variant_stats.items():
            print(f"processing context: {context}")

            if context in self._skip or len(context_stats) > self._skip_limit:
                print(f"skipping context: {context}")
                continue

            self._print_single_context(ws, context, context_stats, start_column)
            start_column = start_column + used_columns + margin

        # change default width
        for col_cells in ws.columns:
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = 12

    def _print_single_context(self, ws, context, context_data, start_column):
        start_row = 2
        used_rows = 0
        margin = 2

        for hash, weight_set_data in context_data.items():
            used_rows = self._print_single_weight_set(ws, weight_set_data, start_column, start_row)
            start_row = start_row + used_rows + margin

    def _print_single_weight_set(self, ws, weight_set_data, start_column, start_row):
        current_column = start_column
        current_row = start_row
        length = 0

        # general information
        ws.merge_cells(start_row=start_row, start_column=start_column, end_row=start_row,end_column=start_column + 2)
        self._add_header(ws, row=start_row, column=start_column, value="General Information")

        self._add_sub_header(ws, row=start_row + 1, column=start_column, value="context")
        ws.merge_cells(start_row=start_row + 1, end_row=start_row + 1, start_column=start_column + 1, end_column=start_column + 2)
        self._add_text(ws, row=start_row + 1, column=start_column + 1, value=weight_set_data["context"])

        self._add_sub_header(ws, row=start_row + 2, column=start_column, value="type")
        ws.merge_cells(start_row=start_row + 2, end_row=start_row + 2, start_column=start_column + 1, end_column=start_column + 2)
        self._add_text(ws, row=start_row + 2, column=start_column + 1, value=weight_set_data["type"])

        self._add_sub_header(ws, row=start_row + 3, column=start_column, value="range")
        ws.merge_cells(start_row=start_row + 3, end_row=start_row + 3, start_column=start_column + 1, end_column=start_column + 2)
        self._add_text(ws, row=start_row + 3, column=start_column + 1, value=weight_set_data["range"])

        self._add_sub_header(ws, row=start_row + 4, column=start_column, value="rng calls")
        ws.merge_cells(start_row=start_row + 4, end_row=start_row + 4, start_column=start_column + 1, end_column=start_column + 2)
        self._add_text(ws, row=start_row + 4, column=start_column + 1, value=weight_set_data["meta"]["numberOfCalls"])

        # run stats (rng)
        ws.merge_cells(start_row=start_row + 5, start_column=start_column, end_row=start_row+5,end_column=start_column + 2)
        self._add_header(ws, row=start_row + 5, column=start_column, value="Run Statistics (rng)")

        self._add_sub_header(ws, row=start_row + 6, column=start_column, value="up")
        self._add_text(ws, row=start_row + 7, column=start_column, value=weight_set_data["runStatsRng"]["up"])

        self._add_sub_header(ws, row=start_row + 6, column=start_column + 1, value="down")
        self._add_text(ws, row=start_row + 7, column=start_column + 1, value=weight_set_data["runStatsRng"]["down"])

        self._add_sub_header(ws, row=start_row + 6, column=start_column + 2, value="same")
        self._add_text(ws, row=start_row + 7, column=start_column + 2, value=weight_set_data["runStatsRng"]["same"])

        # run stats (elements)
        ws.merge_cells(start_row=start_row + 8, start_column=start_column, end_row=start_row+8, end_column=start_column + 2)
        self._add_header(ws, row=start_row + 8, column=start_column, value="Run Statistics (elements)")

        self._add_sub_header(ws, row=start_row + 9, column=start_column, value="up")
        self._add_text(ws, row=start_row + 10, column=start_column, value=weight_set_data["runStatsElements"]["up"])

        self._add_sub_header(ws, row=start_row + 9, column=start_column + 1, value="down")
        self._add_text(ws, row=start_row + 10, column=start_column + 1, value=weight_set_data["runStatsElements"]["down"])

        self._add_sub_header(ws, row=start_row + 9, column=start_column + 2, value="same")
        self._add_text(ws, row=start_row + 10, column=start_column + 2, value=weight_set_data["runStatsElements"]["same"])

        length = length + 11

        # hit table
        hit_table_start_row = start_row + 11
        ws.merge_cells(start_row=hit_table_start_row, start_column=start_column, end_row=hit_table_start_row, end_column=start_column + 2)
        self._add_header(ws, row=hit_table_start_row, column=start_column, value="Hit Table")

        # hit table (has chance)
        if weight_set_data["type"] == "hasChance":
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 0, value="element")
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 1, value="chance")
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 2, value="hits")
            self._add_text(ws, row=hit_table_start_row + 2, column=start_column + 0, value="True")
            self._add_text(ws, row=hit_table_start_row + 2, column=start_column + 1, value=weight_set_data["chance"])
            self._add_text(ws, row=hit_table_start_row + 2, column=start_column + 2, value=weight_set_data["hitTable"]["True"])
            self._add_text(ws, row=hit_table_start_row + 3, column=start_column + 0, value="False")
            self._add_text(ws, row=hit_table_start_row + 3, column=start_column + 1, value=f'1 - {weight_set_data["chance"]}')
            self._add_text(ws, row=hit_table_start_row + 3, column=start_column + 2, value=weight_set_data["hitTable"]["False"])

            length = length + 4

        if weight_set_data["type"] == "randomIndex":
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 0, value="element")
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 1, value="weight")
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 2, value="hits")

            for index, (element, hits) in enumerate(weight_set_data["hitTable"].items()):
                self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 0, value=element)
                self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 1, value="1")
                self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 2, value=hits)

            length = length + 2 + len(weight_set_data["hitTable"])

        if weight_set_data["type"] == "randomWeighted":
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 0, value="element")
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 1, value="weight")
            self._add_sub_header(ws, row=hit_table_start_row + 1, column=start_column + 2, value="hits")

            for index, (element, hits) in enumerate(weight_set_data["hitTable"].items()):
                self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 0, value=element)
                self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 1, value=weight_set_data["weights"][index])
                self._add_text(ws, row=hit_table_start_row + 2 + index, column=start_column + 2, value=hits)

            length = length + 2 + len(weight_set_data["hitTable"])

        # create borders
        thin = Side(style='thin')
        thick = Side(style='thick')
        for row in range(start_row, start_row + length - 1+ 1):
            for col in range(start_column, start_column + 2 + 1):
                cell = ws.cell(row=row, column=col)
                border = Border(
                    left=thick if col == start_column else thin,
                    right=thick if col == start_column + 2 else thin,
                    top=thick if row == start_row else thin,
                    bottom=thick if row == start_row + length - 1 else thin
                )
                cell.border = border

        return length


    def _add_header(self, ws, row, column, value):
        cell = ws.cell(row=row, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
        cell.font = Font(bold=True, color=COLOR_TITLE_TEXT)
        cell.alignment = Alignment(horizontal="center")

    def _add_sub_header(self, ws, row, column, value):
        cell = ws.cell(row=row, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBTITLE_BG)
        cell.font = Font(bold=True, color=COLOR_SUBTITLE_TEXT)
        cell.alignment = Alignment(horizontal="center")

    def _add_text(self, ws, row, column, value):
        cell = ws.cell(row=row, column=column, value=value)
        cell.alignment = Alignment(horizontal="left")


    def save_report_json(self, filename):
        # save report
        os.makedirs("excel_reports", exist_ok=True)
        self._wb.save(filename)


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description="Generate RNG distribution Excel report from a JSON report.")
    parser.add_argument("input", help="Path to the input JSON report file")
    parser.add_argument("output", help="Path for the output Excel (.xlsx) file")
    args = parser.parse_args()

    with open(args.input, 'r') as file:
        data = json.load(file)

    stats_processor = ExcelStatsProcessor()
    stats_processor.set_skip_limit(500)
    stats_processor.create_excel_report(data)
    stats_processor.save_report_json(args.output)
